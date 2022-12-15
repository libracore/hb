# -*- coding: utf-8 -*-
# Copyright (c) 2021, libracore AG and contributors
# For license information, please see license.txt

# import libs
import openpyxl

def get_projects(filename):
    # format definition
    teams = [  # each team section
        {
            'drilling_team': '01',
            'project_row': 8,                # row with the project number
            'customer_name': 9,              # customer name
            'object_name': 10,               # object name
            'object_street': 11,             # object street
            'object_city': 12,               # object plz city and canton
            'object_drilling_detail': 13,    # drilling details
            'object_mud': 14,                # mud disposer MudEx/Fremd
            'object_project_manager': 17     # object project manager (short)
        },
        {
            'drilling_team': '02',
            'project_row': 20,                # row with the project number
            'customer_name': 21,              # customer name
            'object_name': 22,               # object name
            'object_street': 23,             # object street
            'object_city': 24,               # object plz city and canton
            'object_drilling_detail': 25,    # drilling details
            'object_mud': 26,                # mud disposer MudEx/Fremd
            'object_project_manager': 29     # object project manager (short)
        },
        {
            'drilling_team': '03',
            'project_row': 32,                # row with the project number
            'customer_name': 33,              # customer name
            'object_name': 34,               # object name
            'object_street': 35,             # object street
            'object_city': 36,               # object plz city and canton
            'object_drilling_detail': 37,    # drilling details
            'object_mud': 38,                # mud disposer MudEx/Fremd
            'object_project_manager': 41     # object project manager (short)
        },
        {
            'drilling_team': '04',
            'project_row': 44,                # row with the project number
            'customer_name': 45,              # customer name
            'object_name': 46,               # object name
            'object_street': 47,             # object street
            'object_city': 48,               # object plz city and canton
            'object_drilling_detail': 49,    # drilling details
            'object_mud': 50,                # mud disposer MudEx/Fremd
            'object_project_manager': 53     # object project manager (short)
        },
        {
            'drilling_team': '05',
            'project_row': 56,                # row with the project number
            'customer_name': 57,              # customer name
            'object_name': 58,               # object name
            'object_street': 59,             # object street
            'object_city': 60,               # object plz city and canton
            'object_drilling_detail': 61,    # drilling details
            'object_mud': 62,                # mud disposer MudEx/Fremd
            'object_project_manager': 65     # object project manager (short)
        },
        {
            'drilling_team': '06',
            'project_row': 68,                # row with the project number
            'customer_name': 69,              # customer name
            'object_name': 70,               # object name
            'object_street': 71,             # object street
            'object_city': 72,               # object plz city and canton
            'object_drilling_detail': 73,    # drilling details
            'object_mud': 74,                # mud disposer MudEx/Fremd
            'object_project_manager': 77     # object project manager (short)
        },
        {
            'drilling_team': '07',
            'project_row': 80,                # row with the project number
            'customer_name': 81,              # customer name
            'object_name': 82,               # object name
            'object_street': 83,             # object street
            'object_city': 84,               # object plz city and canton
            'object_drilling_detail': 85,    # drilling details
            'object_mud': 86,                # mud disposer MudEx/Fremd
            'object_project_manager': 89     # object project manager (short)
        },
        {
            'drilling_team': '08',
            'project_row': 92,                # row with the project number
            'customer_name': 93,              # customer name
            'object_name': 94,               # object name
            'object_street': 95,             # object street
            'object_city': 96,               # object plz city and canton
            'object_drilling_detail': 97,    # drilling details
            'object_mud': 98,                # mud disposer MudEx/Fremd
            'object_project_manager': 101     # object project manager (short)
        },
        {
            'drilling_team': '09',
            'project_row': 104,                # row with the project number
            'customer_name': 105,              # customer name
            'object_name': 106,               # object name
            'object_street': 107,             # object street
            'object_city': 108,               # object plz city and canton
            'object_drilling_detail': 109,    # drilling details
            'object_mud': 110,                # mud disposer MudEx/Fremd
            'object_project_manager': 113     # object project manager (short)
        },
        {
            'drilling_team': '10',
            'project_row': 116,                # row with the project number
            'customer_name': 117,              # customer name
            'object_name': 118,               # object name
            'object_street': 119,             # object street
            'object_city': 120,               # object plz city and canton
            'object_drilling_detail': 121,    # drilling details
            'object_mud': 122,                # mud disposer MudEx/Fremd
            'object_project_manager': 125     # object project manager (short)
        },
        {
            'drilling_team': '11',
            'project_row': 128,                # row with the project number
            'customer_name': 129,              # customer name
            'object_name': 130,               # object name
            'object_street': 131,             # object street
            'object_city': 132,               # object plz city and canton
            'object_drilling_detail': 133,    # drilling details
            'object_mud': 134,                # mud disposer MudEx/Fremd
            'object_project_manager': 137     # object project manager (short)
        },
        {
            'drilling_team': '12',
            'project_row': 140,                # row with the project number
            'customer_name': 141,              # customer name
            'object_name': 142,               # object name
            'object_street': 143,             # object street
            'object_city': 144,               # object plz city and canton
            'object_drilling_detail': 145,    # drilling details
            'object_mud': 146,                # mud disposer MudEx/Fremd
            'object_project_manager': 149     # object project manager (short)
        },
        {
            'drilling_team': 'Parkplatz',
            'project_row': 152,                # row with the project number
            'customer_name': 153,              # customer name
            'object_name': 154,               # object name
            'object_street': 155,             # object street
            'object_city': 156,               # object plz city and canton
            'object_drilling_detail': 157,    # drilling details
            'object_mud': 158,                # mud disposer MudEx/Fremd
            'object_project_manager': 161     # object project manager (short)
        }
    ]
    date_row = 4            # date field in row 3 (two colums for a day or the weekend)
    first_day = 43          # start reading in column AQ

    # open workbook (use data_only to transfrom formula into data)
    wb_obj = openpyxl.load_workbook(filename, data_only=True)

    # get active sheet
    sheet_obj = wb_obj.active

    # resolve done color
    done_color = sheet_obj['D16'].fill.start_color.index
    print("Done color: {0}".format(done_color))

    # get last used column
    last_column = sheet_obj.max_column

    # loop through teams
    for team in teams:
        print("Reading {0}".format(team['drilling_team']))
        # find first date
        first_project_column = None
        project = None
        current_column = first_day
        while not first_project_column:
            project = sheet_obj.cell(row = team['project_row'], column = current_column).value
            if project:                     # has a project number: this is the first project
                first_project_column = current_column
            current_column += 1             # go to next column
        print("First date: {0}".format(first_project_column))
        print("Project: {0}".format(project))
        start_date = sheet_obj.cell(row = date_row, column = first_project_column).value
        print("Start date from {0}:{1}: {2}".format(date_row, first_project_column, start_date))
        if not start_date:
            start_date = sheet_obj.cell(row = date_row, column = (first_project_column - 1)).value
            start_date_vm = False
        else:
            start_date_vm = True
        projects = [{
            'name': remove_links(project),
            'source': "R{0}:C{1}".format(team['project_row'], first_project_column),
            'start_date': start_date,
            'start_date_vm': start_date_vm,
            'customer_name': sheet_obj.cell(row = team['customer_name'], column = first_project_column).value,
            'object_name': sheet_obj.cell(row = team['object_name'], column = first_project_column).value,
            'object_street': sheet_obj.cell(row = team['object_street'], column = first_project_column).value,
            'object_city': sheet_obj.cell(row = team['object_city'], column = first_project_column).value,
            'object_drilling_detail': sheet_obj.cell(row = team['object_drilling_detail'], column = first_project_column).value,
            'object_mud': sheet_obj.cell(row = team['object_mud'], column = first_project_column).value,
            'object_project_manager': sheet_obj.cell(row = team['object_project_manager'], column = first_project_column).value,
            'status': "Completed" if sheet_obj.cell(row = (team['project_row'] - 1), column = first_project_column).fill.start_color.index == done_color else "Open"
        }]
        # go through columns
        for i in range((first_project_column + 1), last_column):
            project = sheet_obj.cell(row = team['project_row'], column = i).value
            customer_name = sheet_obj.cell(row = team['customer_name'], column = i).value
            # if project and project != projects[-1]['name'] and len(str(project)) == 6:
            if project and project != projects[-1]['name'] and customer_name:
                last_date = sheet_obj.cell(row = date_row, column = (i - 1)).value
                if not last_date:
                    last_date = sheet_obj.cell(row = date_row, column = (i - 2)).value
                    last_date_vm = False
                else:
                    last_date_vm = True
                this_date = sheet_obj.cell(row = date_row, column = i).value
                if not this_date:
                    this_date = sheet_obj.cell(row = date_row, column = (i - 1)).value
                    this_date_vm = False
                else:
                    this_date_vm = True
                projects[-1]['end_date'] = last_date
                projects[-1]['end_date_vm'] = last_date_vm
                projects.append({
                    'name': remove_links(project),
                    'source': "R{0}:C{1}".format(team['project_row'], i),
                    'start_date': this_date,
                    'start_date_vm': this_date_vm,
                    'customer_name': customer_name,
                    'object_name': sheet_obj.cell(row = team['object_name'], column = i).value,
                    'object_street': sheet_obj.cell(row = team['object_street'], column = i).value,
                    'object_city': sheet_obj.cell(row = team['object_city'], column = i).value,
                    'object_drilling_detail': sheet_obj.cell(row = team['object_drilling_detail'], column = i).value,
                    'object_mud': sheet_obj.cell(row = team['object_mud'], column = i).value,
                    'object_project_manager': sheet_obj.cell(row = team['object_project_manager'], column = i).value,
                    'status': "Completed" if sheet_obj.cell(row = (team['project_row'] - 1), column = i).fill.start_color.index == done_color else "Open"
                })

        # add last date
        last_date = sheet_obj.cell(row = date_row, column = last_column).value
        if not last_date:
            last_date = sheet_obj.cell(row = date_row, column = (last_column - 1)).value
            last_date_vm = False
        else:
            last_date_vm = True
        projects[-1]['end_date'] = last_date
        projects[-1]['end_date_vm'] = last_date_vm
        print("{0}".format(projects))
        team['projects'] = projects
    
    # return teams with projects
    return teams

def remove_links(name):
    name = ("{0}".format(name)).replace("\\", "/")
    name = name.split("/")
    return name[-1]
